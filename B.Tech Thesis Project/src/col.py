from openpyxl import Workbook
import matplotlib.pyplot as plt
import pprint
import re
import json
import math
import os
import numpy as np
import peakutils
# ImPortant Function
def convert(input):
    if isinstance(input, dict):
        return {convert(key): convert(value) for key, value in input.iteritems()}
    elif isinstance(input, list):
        return [convert(element) for element in input]
    elif isinstance(input, unicode):
        return input.encode('utf-8')
    else:
        return input


peakmul={}

with open('peakmul_entropyequal.txt','r') as fp:
    peakmulequal=json.loads(fp.read(),encoding='utf-8')
peakmulequal=convert(peakmulequal)
peakmulequal={int(k):[float(i) for i in v] for k,v in peakmulequal.items()}

with open('peakmul_entropy.txt','r') as fp:
    peakmul=json.loads(fp.read(),encoding='utf-8')
peakmul=convert(peakmul)
peakmul={int(k):[float(i) for i in v] for k,v in peakmul.items()}

print('-------------------------------------------------------------------------')

wb = Workbook()
ws = wb.active

i=2
for key in peakmul:
    oo1=peakmul[key]
    oo2=peakmulequal[key]
    ws.cell(row=i,column=1).value=key
    ws.cell(row=i,column=2).value=oo1[0]
    ws.cell(row=i,column=3).value=oo2[0]
    ws.cell(row=i,column=4).value=oo1[1]
    ws.cell(row=i,column=5).value=oo2[1]
    i=i+1
wb.save('Entropy_Peakmul.xlsx')
